"""固定報表產生器 —— 「物件概要書」Excel。

這就是業務端按一個鍵、交到台灣投資人手上的標準提案書：
建物概要 + 投資指標 + レントロール（逐層租金表），中日雙語、日圓 / 台幣雙幣別。
所有物件套用同一份版型，格式永遠一致。
"""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .metrics import compute_metrics, sqm_to_tsubo

# ── 樣式 ──
INK = "111111"
ACCENT = "888888"
CREAM = "F5F0EB"
LINE = "D8D3CE"

TITLE_FONT = Font(name="Yu Gothic", size=20, bold=True, color=INK)
H_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
LABEL_FONT = Font(name="Yu Gothic", size=10, color=ACCENT)
VALUE_FONT = Font(name="Yu Gothic", size=10, color=INK)
BIG_FONT = Font(name="Yu Gothic", size=14, bold=True, color=INK)

HEADER_FILL = PatternFill("solid", fgColor=INK)
BAND_FILL = PatternFill("solid", fgColor=CREAM)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YEN_FMT = '"¥"#,##0'
TWD_FMT = '"NT$"#,##0'
PCT_FMT = '0.00"%"'


def _label_value(ws, row, label, value, value_fmt=None, value_font=VALUE_FONT):
    """寫一組「標籤 / 值」。"""
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = LABEL_FONT
    c1.alignment = Alignment(vertical="center")
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = value_font
    c2.alignment = Alignment(vertical="center")
    if value_fmt:
        c2.number_format = value_fmt
    return row + 1


def _section(ws, row, title_zh, title_en, span=6):
    cell = ws.cell(row=row, column=1, value=f"  {title_zh}　{title_en}")
    cell.font = H_FONT
    cell.alignment = Alignment(vertical="center")
    for col in range(1, span + 1):
        ws.cell(row=row, column=col).fill = HEADER_FILL
    ws.row_dimensions[row].height = 22
    return row + 1


def build_report_workbook(p):
    """回傳一個 BytesIO，內含該物件的「物件概要書」xlsx。"""
    m = compute_metrics(p)
    wb = Workbook()
    ws = wb.active
    ws.title = "物件概要書"
    ws.sheet_view.showGridLines = False

    widths = [22, 26, 16, 12, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 標題 ──
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="物件概要書")
    t.font = TITLE_FONT
    t.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:F2")
    s = ws.cell(row=2, column=1, value="PROPERTY SUMMARY ｜ 日本一棟商辦投資物件")
    s.font = Font(name="Yu Gothic", size=9, color=ACCENT)

    row = 4
    ws.merge_cells(f"A{row}:F{row}")
    nm = ws.cell(row=row, column=1, value=f"{p.code}　{p.name}")
    nm.font = BIG_FONT
    row += 2

    # ── 建物概要 ──
    row = _section(ws, row, "建物概要", "Building")
    loc = "　".join(x for x in [p.prefecture, p.city] if x)
    row = _label_value(ws, row, "所在地  Location", loc or "—")
    row = _label_value(ws, row, "詳細地址  Address", p.address or "—")
    station = f"{p.nearest_station or '—'}　徒步 {p.walk_minutes or '—'} 分"
    row = _label_value(ws, row, "最寄り駅  Station", station)
    floors_txt = f"地上 {p.floors_above or 0} 階 / 地下 {p.floors_below or 0} 階"
    row = _label_value(ws, row, "規模  Floors", floors_txt)
    row = _label_value(ws, row, "構造  Structure", p.structure or "—")
    built = f"{p.built_year or '—'} 年 {p.built_month or ''} 月".strip()
    row = _label_value(ws, row, "築年  Built", built)
    row = _label_value(ws, row, "土地面積  Land", f"{p.land_area_sqm or 0:,.2f} ㎡（{sqm_to_tsubo(p.land_area_sqm)} 坪）")
    row = _label_value(ws, row, "延床面積  GFA", f"{p.gross_floor_area_sqm or 0:,.2f} ㎡（{sqm_to_tsubo(p.gross_floor_area_sqm)} 坪）")
    row = _label_value(ws, row, "權利型態  Title", p.ownership_type or "—")
    row = _label_value(ws, row, "耐震基準  Seismic", p.earthquake_standard or "—")
    row += 1

    # ── 投資指標 ──
    row = _section(ws, row, "投資指標", "Investment")
    row = _label_value(ws, row, "販售價格  Price (JPY)", p.price_jpy, YEN_FMT, BIG_FONT)
    row = _label_value(ws, row, "販售價格  Price (TWD)", m["price_twd"], TWD_FMT, BIG_FONT)
    row = _label_value(ws, row, "參考匯率  FX (1 JPY)", f'NT$ {m["exchange_rate"]}')
    row = _label_value(ws, row, "表面利回り  Gross Yield", m["gross_yield"], PCT_FMT, BIG_FONT)
    row = _label_value(ws, row, "現況利回り  Current Yield", m["current_yield"], PCT_FMT)
    row = _label_value(ws, row, "滿室想定年收  Income (full)", m["full_annual_income_jpy"], YEN_FMT)
    row = _label_value(ws, row, "現況年收  Income (current)", m["current_annual_income_jpy"], YEN_FMT)
    row = _label_value(ws, row, "入居率  Occupancy", m["occupancy_rate"], PCT_FMT)
    row = _label_value(ws, row, "固定資產稅  Tax / yr", p.property_tax_jpy or 0, YEN_FMT)
    row = _label_value(ws, row, "管理修繕費  Mgmt / yr", p.management_fee_jpy or 0, YEN_FMT)
    row += 1

    # ── レントロール ──
    row = _section(ws, row, "レントロール", "Rent Roll")
    headers = ["樓層 / 區劃", "面積(坪)", "現況", "租客 / 業種", "月租金(¥)", "共益費(¥)"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(name="Yu Gothic", size=9, bold=True, color=INK)
        c.fill = BAND_FILL
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for f in p.floors:
        unit = f"{f.floor_label or ''}{('  ' + f.unit) if f.unit else ''}".strip()
        tenant = f.tenant_name or ("—" if f.status != "賃貸中" else "")
        if f.tenant_industry:
            tenant = f"{tenant}（{f.tenant_industry}）"
        cells = [
            unit,
            sqm_to_tsubo(f.area_sqm),
            f.status,
            tenant,
            f.monthly_rent_jpy or 0,
            f.monthly_common_fee_jpy or 0,
        ]
        for col, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Yu Gothic", size=9, color=INK)
            c.border = BORDER
            if col in (5, 6):
                c.number_format = YEN_FMT
            if col in (2, 3):
                c.alignment = Alignment(horizontal="center")
        if f.status == "空室":
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor="FBEAEA")
        row += 1

    # 合計列
    c = ws.cell(row=row, column=1, value="合計  Total")
    c.font = Font(name="Yu Gothic", size=9, bold=True, color=INK)
    c.fill = BAND_FILL
    c.border = BORDER
    ws.cell(row=row, column=2, value=m["total_area_tsubo"]).font = Font(name="Yu Gothic", size=9, bold=True)
    ws.cell(row=row, column=2).fill = BAND_FILL
    ws.cell(row=row, column=2).border = BORDER
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
    tot = ws.cell(row=row, column=5, value=m["full_monthly_income_jpy"])
    tot.font = Font(name="Yu Gothic", size=9, bold=True)
    tot.number_format = YEN_FMT
    tot.fill = BAND_FILL
    tot.border = BORDER
    for col in (3, 4, 6):
        cc = ws.cell(row=row, column=col)
        cc.fill = BAND_FILL
        cc.border = BORDER
    row += 2

    # ── 頁尾 ──
    ws.merge_cells(f"A{row}:F{row}")
    foot = ws.cell(row=row, column=1,
                   value=f"製表日：{date.today():%Y-%m-%d}　｜　負責業務：{p.assigned_sales or '—'}　｜　銷售狀態：{p.status}")
    foot.font = Font(name="Yu Gothic", size=9, color=ACCENT)
    row += 1
    ws.merge_cells(f"A{row}:F{row}")
    disc = ws.cell(row=row, column=1,
                   value="※ 本資料僅供參考，台幣金額依參考匯率換算，實際成交以日圓計價為準。利回り為稅前數字。")
    disc.font = Font(name="Yu Gothic", size=8, color=ACCENT)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
