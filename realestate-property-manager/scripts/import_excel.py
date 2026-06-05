"""舊 Excel 轉檔工具 —— 把「一物件一 Excel」批次灌進資料庫。

用法：
    python scripts/import_excel.py scripts/sample_legacy

讀取資料夾內每一個 .xlsx，依下列版型解析後寫入資料庫
（同 code 已存在則更新，達成可重複執行的搬遷）：

  ┌ 工作表「物件概要」：A 欄為欄位名、B 欄為值（label → value）
  └ 工作表「レントロール」：第一列為表頭，其後每列一筆樓層

面積一律以「坪」輸入，匯入時自動換算成 ㎡ 儲存。
真實專案中，只要照客戶各自的 Excel 調整 LABEL_MAP / 表頭對照即可。
"""
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from app.database import Base, SessionLocal, engine  # noqa: E402
from app.models import Floor, Property  # noqa: E402

TSUBO = 3.305785

# 物件概要表：Excel 標籤 → 資料庫欄位
LABEL_MAP = {
    "物件編號": "code", "物件名稱": "name",
    "都道府県": "prefecture", "市區町村": "city", "詳細地址": "address",
    "最寄り駅": "nearest_station", "徒步分": "walk_minutes",
    "構造": "structure", "地上階": "floors_above", "地下階": "floors_below",
    "築年": "built_year", "築月": "built_month",
    "土地面積㎡": "land_area_sqm", "延床面積㎡": "gross_floor_area_sqm",
    "權利型態": "ownership_type", "耐震": "earthquake_standard",
    "販售價格日圓": "price_jpy", "匯率": "exchange_rate",
    "固定資產稅": "property_tax_jpy", "管理修繕費": "management_fee_jpy",
    "銷售狀態": "status", "負責業務": "assigned_sales", "備註": "note",
}
INT_FIELDS = {"walk_minutes", "floors_above", "floors_below", "built_year", "built_month"}
FLOAT_FIELDS = {"land_area_sqm", "gross_floor_area_sqm", "price_jpy", "exchange_rate",
                "property_tax_jpy", "management_fee_jpy"}

# レントロール表頭 → 樓層欄位（面積為坪）
RENT_HEADERS = {
    "樓層": "floor_label", "區劃": "unit", "面積(坪)": "area_tsubo", "現況": "status",
    "租客": "tenant_name", "業種": "tenant_industry", "月租金": "monthly_rent_jpy",
    "共益費": "monthly_common_fee_jpy", "想定月租": "assumed_monthly_rent_jpy",
    "押金": "deposit_jpy", "契約開始": "lease_start", "契約結束": "lease_end", "備註": "note",
}


def _cast(field, value):
    if value in (None, ""):
        return None
    if field in INT_FIELDS:
        return int(float(value))
    if field in FLOAT_FIELDS:
        return float(value)
    return str(value).strip()


def parse_property_sheet(ws):
    data = {}
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = (row + (None, None))[:2]
        if label and str(label).strip() in LABEL_MAP:
            field = LABEL_MAP[str(label).strip()]
            data[field] = _cast(field, value)
    return data


def parse_rent_roll(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    idx = {RENT_HEADERS[h]: i for i, h in enumerate(header) if h in RENT_HEADERS}
    floors = []
    for order, raw in enumerate(rows[1:], start=1):
        if not any(raw):
            continue
        get = lambda k: raw[idx[k]] if k in idx and idx[k] < len(raw) else None
        tsubo = get("area_tsubo") or 0
        floors.append(Floor(
            sort_order=order,
            floor_label=str(get("floor_label") or "").strip(),
            unit=(str(get("unit")).strip() if get("unit") else None),
            area_sqm=round(float(tsubo) * TSUBO, 2),
            status=str(get("status") or "空室").strip(),
            tenant_name=(str(get("tenant_name")).strip() if get("tenant_name") else None),
            tenant_industry=(str(get("tenant_industry")).strip() if get("tenant_industry") else None),
            monthly_rent_jpy=float(get("monthly_rent_jpy") or 0),
            monthly_common_fee_jpy=float(get("monthly_common_fee_jpy") or 0),
            assumed_monthly_rent_jpy=float(get("assumed_monthly_rent_jpy") or 0),
            deposit_jpy=float(get("deposit_jpy") or 0),
            lease_start=(str(get("lease_start")).strip() if get("lease_start") else None),
            lease_end=(str(get("lease_end")).strip() if get("lease_end") else None),
        ))
    return floors


def import_file(path, db):
    wb = load_workbook(path, data_only=True)
    if "物件概要" not in wb.sheetnames:
        print(f"  ⚠ 略過 {path.name}：找不到『物件概要』工作表")
        return False
    data = parse_property_sheet(wb["物件概要"])
    if not data.get("code"):
        print(f"  ⚠ 略過 {path.name}：缺少物件編號")
        return False
    floors = parse_rent_roll(wb["レントロール"]) if "レントロール" in wb.sheetnames else []

    existing = db.query(Property).filter_by(code=data["code"]).first()
    if existing:
        for k, v in data.items():
            setattr(existing, k, v)
        existing.floors.clear()
        existing.floors.extend(floors)
        verb = "更新"
    else:
        db.add(Property(**data, floors=floors))
        verb = "新增"
    print(f"  ✓ {verb} {data['code']} {data.get('name','')}（{len(floors)} 層）")
    return True


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("scripts/sample_legacy")
    files = sorted(folder.glob("*.xlsx"))
    if not files:
        print(f"在 {folder} 找不到任何 .xlsx 檔。")
        return
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        print(f"開始匯入 {len(files)} 個檔案（來源：{folder}）")
        ok = sum(import_file(p, db) for p in files)
        db.commit()
        print(f"完成：成功匯入 {ok} / {len(files)} 個物件。")
    finally:
        db.close()


if __name__ == "__main__":
    main()
